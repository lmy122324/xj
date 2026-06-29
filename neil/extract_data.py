"""
Neil 期货分析看板 - 数据提取脚本 v2
精准解析 PVC数据表.xlsx 的"图表"sheet 结构
"""
import json, os, re
from datetime import datetime
import openpyxl

SRC = r'C:\Users\19916\Desktop\期货分析数据库\PVC数据表.xlsx'
OUT = os.path.join(os.path.dirname(__file__), 'data.json')

def safe_float(v):
    if v is None: return None
    if isinstance(v, (int, float)): return round(float(v), 2)
    s = str(v).replace(',', '').replace('%', '').strip()
    # Handle "41.54万吨" format
    m = re.match(r'([\d.]+)', s)
    if m: return round(float(m.group(1)), 2)
    return None

def parse_mom_yoy(v):
    """Parse '0.45/1.09%' → (value, mom, yoy)"""
    if v is None: return (None, None, None)
    s = str(v)
    parts = s.split('/')
    val = safe_float(parts[0]) if len(parts) > 0 else None
    mom = safe_float(parts[0]) if len(parts) > 0 else None
    yoy = safe_float(parts[1]) if len(parts) > 1 else None
    return (val, mom, yoy)

def extract_all(wb):
    ws = wb['图表']
    
    data = {
        'update_time': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'latest_week': {},
        'weekly_trend': [],
        'price_table': [],
        'inventory_trend': [],
        'supply_demand': {},
    }
    
    # === Section 1: 周度核心数据 (R40-R48, C1-4) ===
    weekly_metrics = {}
    for r in range(40, 56):
        label = ws.cell(row=r, column=1).value
        val = ws.cell(row=r, column=2).value
        mom_yoy = ws.cell(row=r, column=3).value
        if label and val:
            label = str(label).strip()
            v, mom, yoy = parse_mom_yoy(val) if '/' in str(val) else (safe_float(val), None, None)
            weekly_metrics[label] = {
                'value': v,
                'mom': mom if mom else parse_mom_yoy(mom_yoy)[0],
                'yoy': yoy if yoy else parse_mom_yoy(mom_yoy)[2],
            }
    
    data['latest_week'] = weekly_metrics
    
    # === Section 2: 价格/价差/成本/利润表 (R17-R34, C1-9) ===
    # C1=项目类型, C2=子项, C4=数值
    price_rows = []
    current_category = ''
    for r in range(17, 35):
        cat = ws.cell(row=r, column=1).value
        name = ws.cell(row=r, column=2).value
        val_c4 = ws.cell(row=r, column=4).value
        val_c8 = ws.cell(row=r, column=8).value
        
        if cat and str(cat).strip():
            current_category = str(cat).strip()
        
        row_data = {
            'category': current_category,
            'name': str(name).strip() if name else '',
            'value_left': safe_float(val_c4),
            'value_right': safe_float(val_c8),
        }
        if row_data['name']:
            price_rows.append(row_data)
    
    data['price_table'] = price_rows
    
    # === Section 3: PVC供应+库存 (R2-R11, C23-27) ===
    supply_data = {}
    for r in range(1, 12):
        label = ws.cell(row=r, column=23).value
        name = ws.cell(row=r, column=24).value
        val = ws.cell(row=r, column=25).value
        mom = ws.cell(row=r, column=26).value
        yoy = ws.cell(row=r, column=27).value
        
        if label:
            current_section = str(label).strip()
        
        if name and val:
            key = str(name).strip()
            if current_section not in supply_data:
                supply_data[current_section] = {}
            supply_data[current_section][key] = {
                'value': safe_float(val) if not isinstance(val, str) else str(val),
                'mom': str(mom) if mom else None,
                'yoy': str(yoy) if yoy else None,
            }
    
    data['supply_demand'] = supply_data
    
    # === Section 4: 历年库存趋势 (R8-R16, C43-47) ===
    inventory = []
    for r in range(8, 17):
        year = ws.cell(row=r, column=43).value
        inv_type = ws.cell(row=r, column=44).value
        v1 = ws.cell(row=r, column=45).value
        v2 = ws.cell(row=r, column=46).value
        v3 = ws.cell(row=r, column=47).value
        
        if year and inv_type:
            inventory.append({
                'year': str(year).strip(),
                'type': str(inv_type).strip(),
                'val1': safe_float(v1),
                'val2': safe_float(v2),
                'val3': safe_float(v3),
            })
    
    data['inventory_trend'] = inventory
    
    # === Section 5: 周度sheet数据 ===
    ws2 = wb['周度']
    weekly_trend = []
    for r in range(1, min(200, ws2.max_row + 1)):
        vals = []
        for c in range(1, 10):
            v = ws2.cell(row=r, column=c).value
            if isinstance(v, datetime):
                vals.append(('date', v.strftime('%Y-%m-%d')))
            elif v is not None:
                vals.append(('val', safe_float(v)))
        if vals:
            weekly_trend.append(vals)
    
    data['weekly_raw'] = weekly_trend
    
    return data

def main():
    print('Loading Excel...')
    wb = openpyxl.load_workbook(SRC, data_only=True)
    
    print('Extracting data...')
    data = extract_all(wb)
    wb.close()
    
    with open(OUT, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)
    
    print(f'Done → {OUT}')
    print(f'  latest_week: {len(data["latest_week"])} metrics')
    print(f'  price_table: {len(data["price_table"])} rows')
    print(f'  supply_demand: {len(data["supply_demand"])} sections')
    print(f'  inventory_trend: {len(data["inventory_trend"])} records')
    
    # Print key metrics
    print('\n=== 最新周度核心数据 ===')
    for k, v in data['latest_week'].items():
        print(f'  {k}: {v}')

if __name__ == '__main__':
    main()
