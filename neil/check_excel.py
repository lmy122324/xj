import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\19916\Desktop\期货分析数据库\PVC数据表.xlsx', data_only=True)
ws = wb['图表']
print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')

for row_idx in range(1, 50):
    row_data = []
    for col_idx in range(1, 50):
        cell = ws.cell(row=row_idx, column=col_idx)
        v = cell.value
        if v is not None:
            row_data.append(f'C{col_idx}={v}')
    if row_data:
        print(f'R{row_idx}: ' + ' | '.join(row_data[:12]))

wb.close()
